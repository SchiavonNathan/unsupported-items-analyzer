#!/usr/bin/env python3
"""Generate a report of Zabbix items affected by collection-related problems.

The script uses pyzabbix and reads configuration from environment variables,
optionally loaded from a .env file.

Supported authentication modes:
- ZABBIX_TOKEN: preferred. The script injects the token into the pyzabbix client
  without performing a login.
- ZABBIX_USER / ZABBIX_PASSWORD: fallback for environments that still use
  username/password authentication.

The report focuses on active trigger problems and keeps the default scope on
collection-related issues such as nodata(), while allowing additional keywords
to be configured.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - optional runtime dependency
    def load_dotenv() -> None:
        return None

from pyzabbix import ZabbixAPI


SEVERITY_LABELS = {
    0: "Not classified",
    1: "Information",
    2: "Warning",
    3: "Average",
    4: "High",
    5: "Disaster",
}


@dataclass(frozen=True)
class Settings:
    zabbix_url: str
    zabbix_token: Optional[str]
    zabbix_user: Optional[str]
    zabbix_password: Optional[str]
    cache_ttl: int
    keywords: List[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a report of Zabbix items with collection problems."
    )
    parser.add_argument(
        "--days",
        type=int,
        default=0,
        help="Limit the search to problems created in the last N days. 0 means all active problems.",
    )
    parser.add_argument(
        "--keywords",
        default="",
        help=(
            "Comma-separated keywords used to catch collection-related problems "
            "besides nodata(). Overrides COLLECTION_KEYWORDS if provided."
        ),
    )
    parser.add_argument(
        "--report",
        choices=("unsupported", "masters", "priorities"),
        default="unsupported",
        help=(
            "Which report to generate. 'unsupported' (default) lists items in "
            "unsupported state. 'masters' lists master collection items and "
            "whether they have a nodata() failure alert. 'priorities' groups "
            "failures by error pattern into a Pareto-ranked resolution order."
        ),
    )
    parser.add_argument(
        "--format",
        choices=("table", "json", "csv", "excel"),
        default="table",
        help="Output format.",
    )
    parser.add_argument(
        "--output",
        default="",
        help="Optional output file for CSV, JSON or Excel formats. If omitted, prints to stdout.",
    )
    parser.add_argument(
        "--serve",
        action="store_true",
        help="Run a local Flask web server to show reports on a premium frontend.",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=5000,
        help="Port to run the Flask web server on (default: 5000).",
    )
    return parser.parse_args()


def load_settings(args: argparse.Namespace) -> Settings:
    load_dotenv()

    zabbix_url = os.getenv("ZABBIX_URL", "").strip()
    if not zabbix_url:
        raise SystemExit("ZABBIX_URL is required in the environment or .env file.")

    token = os.getenv("ZABBIX_TOKEN", "").strip() or None
    user = os.getenv("ZABBIX_USER", "").strip() or None
    password = os.getenv("ZABBIX_PASSWORD", "").strip() or None
    cache_ttl = int(os.getenv("CACHE_TTL", "300"))

    keyword_source = args.keywords.strip()
    if keyword_source:
        keywords = [item.strip().lower() for item in keyword_source.split(",") if item.strip()]
    else:
        env_keywords = os.getenv("COLLECTION_KEYWORDS", "nodata,no data,not supported,unsupported")
        keywords = [item.strip().lower() for item in env_keywords.split(",") if item.strip()]

    return Settings(
        zabbix_url=zabbix_url,
        zabbix_token=token,
        zabbix_user=user,
        zabbix_password=password,
        cache_ttl=cache_ttl,
        keywords=keywords,
    )


def normalize_base_url(url: str) -> str:
    return url.rstrip("/")


def build_api(settings: Settings) -> ZabbixAPI:
    base_url = normalize_base_url(settings.zabbix_url)

    if settings.zabbix_token:
        # Standard instantiation ensures compatibility with both py-zabbix and pyzabbix libraries
        api = ZabbixAPI(base_url)
        api.auth = settings.zabbix_token
        return api

    if not settings.zabbix_user or not settings.zabbix_password:
        raise SystemExit(
            "Provide either ZABBIX_TOKEN or both ZABBIX_USER and ZABBIX_PASSWORD."
        )

    return ZabbixAPI(
        url=base_url,
        user=settings.zabbix_user,
        password=settings.zabbix_password,
    )


ITEM_TYPES = {
    "0": "Zabbix agent",
    "1": "SNMPv1 agent",
    "2": "Zabbix trapper",
    "3": "Simple check",
    "4": "SNMPv2 agent",
    "5": "Zabbix internal",
    "6": "SNMPv3 agent",
    "7": "Zabbix agent (active)",
    "8": "Zabbix aggregate",
    "9": "Web item",
    "10": "External check",
    "11": "Database monitor",
    "12": "IPMI agent",
    "13": "SSH agent",
    "14": "Telnet agent",
    "15": "Calculated",
    "16": "JMX agent",
    "17": "SNMP trap",
    "18": "Dependent item",
    "19": "HTTP agent",
    "20": "SNMP agent",
    "21": "Script",
}


def get_item_type_label(type_code: Any) -> str:
    return ITEM_TYPES.get(str(type_code), f"Other ({type_code})")


# Item types that act as "master collection" sources: they typically collect a
# bulk payload (e.g. a JSON) that is then split across dependent items via
# preprocessing. We only treat them as master collections when they actually
# have dependent items (see collect_master_items).
MASTER_COLLECTION_TYPES = {
    "10": "External check",
    "11": "Database monitor",
    "13": "SSH agent",
    "14": "Telnet agent",
    "15": "Calculated",
    "19": "HTTP agent",
    "21": "Script",
}

# Dependent item type code in the Zabbix API.
DEPENDENT_ITEM_TYPE = "18"

# Substring used to detect nodata()-based collection-failure triggers.
NODATA_KEYWORD = "nodata"


def render_table(rows: Sequence[Dict[str, Any]]) -> str:
    if not rows:
        return "Nenhum item não suportado encontrado."

    headers = ["Host", "Item", "Key", "Type", "Error"]
    columns = [
        [row["host"] for row in rows],
        [row["item"] for row in rows],
        [row["key"] for row in rows],
        [row["type"] for row in rows],
        [row["error"] for row in rows],
    ]
    widths = [
        max(len(headers[index]), *(len(str(value)) for value in column))
        for index, column in enumerate(columns)
    ]

    def fmt_row(values: Sequence[str]) -> str:
        return " | ".join(str(value).ljust(widths[index]) for index, value in enumerate(values))

    lines = [fmt_row(headers), "-+-".join("-" * width for width in widths)]
    for row in rows:
        lines.append(fmt_row([row["host"], row["item"], row["key"], row["type"], row["error"]]))

    return "\n".join(lines)


def render_masters_table(rows: Sequence[Dict[str, Any]]) -> str:
    if not rows:
        return "Nenhum item master de coleta encontrado."

    headers = ["Template", "Item", "Key", "Type", "Dep.", "Alerta nodata", "Triggers nodata"]
    columns = [
        [row["template"] for row in rows],
        [row["item"] for row in rows],
        [row["key"] for row in rows],
        [row["type"] for row in rows],
        [str(row["dependents"]) for row in rows],
        ["SIM" if row["has_nodata_alert"] else "FALTA" for row in rows],
        [row["nodata_triggers"] for row in rows],
    ]
    widths = [
        max(len(headers[index]), *(len(str(value)) for value in column))
        for index, column in enumerate(columns)
    ]

    def fmt_row(values: Sequence[str]) -> str:
        return " | ".join(str(value).ljust(widths[index]) for index, value in enumerate(values))

    lines = [fmt_row(headers), "-+-".join("-" * width for width in widths)]
    for row in rows:
        lines.append(fmt_row([
            row["template"],
            row["item"],
            row["key"],
            row["type"],
            str(row["dependents"]),
            "SIM" if row["has_nodata_alert"] else "FALTA",
            row["nodata_triggers"],
        ]))

    return "\n".join(lines)


def render_priorities_table(rows: Sequence[Dict[str, Any]]) -> str:
    if not rows:
        return "Nenhuma falha encontrada para priorizar."

    headers = ["#", "Prioridade", "Padrão de Erro", "Itens", "Hosts", "% Acum.", "Tipo", "Categoria"]
    columns = [
        [str(row["rank"]) for row in rows],
        [row["priority"] for row in rows],
        [row["pattern"] for row in rows],
        [str(row["count"]) for row in rows],
        [str(row["hosts"]) for row in rows],
        [f"{row['cum_pct'] * 100:.1f}%" for row in rows],
        [row["top_type"] for row in rows],
        [row["top_category"] for row in rows],
    ]
    widths = [
        max(len(headers[index]), *(len(str(value)) for value in column))
        for index, column in enumerate(columns)
    ]

    def fmt_row(values: Sequence[str]) -> str:
        return " | ".join(str(value).ljust(widths[index]) for index, value in enumerate(values))

    lines = [fmt_row(headers), "-+-".join("-" * width for width in widths)]
    for row in rows:
        lines.append(fmt_row([
            str(row["rank"]),
            row["priority"],
            row["pattern"],
            str(row["count"]),
            str(row["hosts"]),
            f"{row['cum_pct'] * 100:.1f}%",
            row["top_type"],
            row["top_category"],
        ]))

    return "\n".join(lines)


def normalize_error(error_msg: str) -> str:
    import re
    if not error_msg:
        return "Unsupported item (no error description)"
    
    # Replace hex hashes or long alphanumerics (length >= 8) with <hash>
    msg = re.sub(r'\b[a-fA-F0-9]{8,}\b', '<hash>', error_msg)
    # Replace interface patterns like cali...
    msg = re.sub(r'\bcali[a-zA-Z0-9]+\b', 'cali<id>', msg)
    # Replace numbers
    msg = re.sub(r'\b\d+\b', '#', msg)
    # Replace single/double quoted strings (dynamic DB names, etc.)
    msg = re.sub(r"\'[^\']+\'", "'<string>'", msg)
    msg = re.sub(r'\"[^\"]+\"', '"<string>"', msg)
    # Normalize extra spacing
    msg = re.sub(r'\s+', ' ', msg).strip()
    return msg


def classify_host(group_names: List[str]) -> str:
    import re
    
    # Precedence patterns as requested
    patterns = [
        ("Databases", re.compile(r"Databases.*|Oracle RAC|.*SQL.*|Templates/Databases", re.IGNORECASE)),
        ("Network", re.compile(r"Network.*", re.IGNORECASE)),
        ("Applications", re.compile(r"CadeiaCTE.*|VLI SIOP|VLI UNICOM|VLI UNILOG|.*Middleware.*|.*Web Monitoring.*|.*Terminal.*|.*Tibco.*", re.IGNORECASE)),
        ("Infrastructure", re.compile(r".*Commvault.*|.*vCenter.*|.*Nobreaks.*|.*SECINFO.*|Zabbix servers.*|.*SYNCRO.*|templates-teste", re.IGNORECASE)),
        ("Servers & Cloud", re.compile(r".*Linux.*|.*Windows.*|.*Cloud.*|.*VMware|Oracle Cloud Infrastructure.*", re.IGNORECASE)),
    ]
    
    for group in group_names:
        for cat_name, regex in patterns:
            if regex.search(group):
                return cat_name
                
    return "All"


def generate_excel_in_memory(rows: Sequence[Dict[str, Any]]) -> io.BytesIO:
    import io
    from collections import Counter, defaultdict
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Stylings
    title_font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. Sheet: Resumo
    ws_sum = wb.active
    ws_sum.title = "Resumo"
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Title
    ws_sum.merge_cells("A1:E2")
    ws_sum["A1"] = "Diagnóstico de Falhas de Coleta Zabbix"
    ws_sum["A1"].font = title_font
    ws_sum["A1"].fill = blue_fill
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws_sum["A4"] = "Métrica Geral"
    ws_sum["B4"] = "Quantidade"
    ws_sum["A4"].font = header_font
    ws_sum["B4"].font = header_font
    ws_sum["A4"].fill = blue_fill
    ws_sum["B4"].fill = blue_fill
    
    total_items = len(rows)
    unique_hosts = len(set(r["host"] for r in rows if r["host"] != "-"))
    
    ws_sum["A5"] = "Total de Itens Não Suportados"
    ws_sum["B5"] = total_items
    ws_sum["A6"] = "Total de Hosts com Erros"
    ws_sum["B6"] = unique_hosts
    
    for row in (5, 6):
        ws_sum[f"A{row}"].font = bold_font
        ws_sum[f"B{row}"].font = regular_font
        ws_sum[f"B{row}"].number_format = "#,##0"
        ws_sum[f"A{row}"].border = thin_border
        ws_sum[f"B{row}"].border = thin_border

    # Table: Maiores Ofensores por Categoria
    ws_sum["A8"] = "Categoria de Grupo de Host"
    ws_sum["B8"] = "Qtd. Falhas"
    ws_sum["C8"] = "% do Total"
    ws_sum["D8"] = "Maior Ofensor (Host)"
    ws_sum["E8"] = "Erro Principal da Categoria"
    
    for col in ("A8", "B8", "C8", "D8", "E8"):
        ws_sum[col].font = header_font
        ws_sum[col].fill = blue_fill
        ws_sum[col].alignment = Alignment(horizontal="left" if col in ("A8", "D8", "E8") else "right")
        
    category_data = defaultdict(list)
    for r in rows:
        category_data[r["category"]].append(r)
        
    categories_to_show = ["Databases", "Network", "Applications", "Infrastructure", "Servers & Cloud", "All"]
    
    current_row = 9
    for idx, cat_name in enumerate(categories_to_show):
        cat_items = category_data.get(cat_name, [])
        cat_count = len(cat_items)
        pct = cat_count / total_items if total_items > 0 else 0
        
        if cat_items:
            host_counts = Counter(item["host"] for item in cat_items)
            top_host, top_host_count = host_counts.most_common(1)[0]
            offender_str = f"{top_host} ({top_host_count} falhas)"
            
            norm_errors = [normalize_error(item["error"]) for item in cat_items]
            top_error = Counter(norm_errors).most_common(1)[0][0]
        else:
            offender_str = "-"
            top_error = "-"
            
        ws_sum[f"A{current_row}"] = cat_name
        ws_sum[f"B{current_row}"] = cat_count
        ws_sum[f"C{current_row}"] = pct
        ws_sum[f"D{current_row}"] = offender_str
        ws_sum[f"E{current_row}"] = top_error
        
        ws_sum[f"B{current_row}"].number_format = "#,##0"
        ws_sum[f"C{current_row}"].number_format = "0.0%"
        
        use_zebra = (idx % 2 == 1)
        for col in ("A", "B", "C", "D", "E"):
            cell = ws_sum[f"{col}{current_row}"]
            cell.font = regular_font
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill
            if col in ("B", "C"):
                cell.alignment = Alignment(horizontal="right")
        current_row += 1

    # Table Type Distribution (starts at row 17)
    ws_sum["A17"] = "Tipo de Item (Protocolo)"
    ws_sum["B17"] = "Quantidade de Falhas"
    ws_sum["C17"] = "% do Total"
    
    for col in ("A17", "B17", "C17"):
        ws_sum[col].font = header_font
        ws_sum[col].fill = blue_fill
        ws_sum[col].alignment = Alignment(horizontal="left" if col == "A17" else "right")
        
    type_counts = Counter(r["type"] for r in rows)
    current_row = 18
    for idx, (item_type, count) in enumerate(type_counts.most_common()):
        pct = count / total_items if total_items > 0 else 0
        ws_sum[f"A{current_row}"] = item_type
        ws_sum[f"B{current_row}"] = count
        ws_sum[f"C{current_row}"] = pct
        
        ws_sum[f"B{current_row}"].number_format = "#,##0"
        ws_sum[f"C{current_row}"].number_format = "0.0%"
        
        use_zebra = (idx % 2 == 1)
        for col in ("A", "B", "C"):
            cell = ws_sum[f"{col}{current_row}"]
            cell.font = regular_font
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill
            if col != "A":
                cell.alignment = Alignment(horizontal="right")
        current_row += 1

    # 2. Sheet: Padroes_de_Erro
    ws_pat = wb.create_sheet(title="Padroes_de_Erro")
    ws_pat.views.sheetView[0].showGridLines = True
    
    ws_pat.merge_cells("A1:C2")
    ws_pat["A1"] = "Padrões de Erros Mais Frequentes"
    ws_pat["A1"].font = title_font
    ws_pat["A1"].fill = blue_fill
    ws_pat["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws_pat["A4"] = "Padrão de Mensagem de Erro (Normalizada)"
    ws_pat["B4"] = "Quantidade"
    ws_pat["C4"] = "% do Total"
    
    for col in ("A4", "B4", "C4"):
        ws_pat[col].font = header_font
        ws_pat[col].fill = blue_fill
        ws_pat[col].alignment = Alignment(horizontal="left" if col == "A4" else "right")
        
    normalized_errors = [normalize_error(r["error"]) for r in rows]
    pattern_counts = Counter(normalized_errors)
    
    current_row = 5
    for idx, (pattern, count) in enumerate(pattern_counts.most_common(50)):
        pct = count / total_items if total_items > 0 else 0
        ws_pat[f"A{current_row}"] = pattern
        ws_pat[f"B{current_row}"] = count
        ws_pat[f"C{current_row}"] = pct
        
        ws_pat[f"B{current_row}"].number_format = "#,##0"
        ws_pat[f"C{current_row}"].number_format = "0.0%"
        
        use_zebra = (idx % 2 == 1)
        for col in ("A", "B", "C"):
            cell = ws_pat[f"{col}{current_row}"]
            cell.font = regular_font
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill
            if col != "A":
                cell.alignment = Alignment(horizontal="right")
        current_row += 1

    # 3. Sheet: Itens_Detalhados
    ws_det = wb.create_sheet(title="Itens_Detalhados")
    ws_det.views.sheetView[0].showGridLines = True
    
    headers = ["Host", "ID do Host", "Item", "Chave (Key)", "Tipo de Coleta", "Categoria de Grupo", "Mensagem de Erro", "ID do Item"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="left")
        
    for idx, r in enumerate(rows):
        row_idx = idx + 2
        row_data = [
            r["host"],
            r["host_id"],
            r["item"],
            r["key"],
            r["type"],
            r["category"],
            r["error"],
            r["item_id"]
        ]
        
        use_zebra = (idx % 2 == 1)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_det.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill

    # Auto-adjust column widths
    for ws in (ws_sum, ws_pat, ws_det):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                if ws.title in ("Resumo", "Padroes_de_Erro") and cell.row <= 2:
                    continue
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            
            width = min(max(max_len + 3, 10), 65)
            ws.column_dimensions[col_letter].width = width

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_masters_excel_in_memory(rows: Sequence[Dict[str, Any]]) -> "io.BytesIO":
    import io
    from collections import Counter
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    title_font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)

    blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    total = len(rows)
    with_alert = sum(1 for r in rows if r["has_nodata_alert"])
    without_alert = total - with_alert

    # 1. Sheet: Resumo
    ws_sum = wb.active
    ws_sum.title = "Resumo"
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum.merge_cells("A1:C2")
    ws_sum["A1"] = "Cobertura de Alertas nodata() em Coletas Master"
    ws_sum["A1"].font = title_font
    ws_sum["A1"].fill = blue_fill
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum["A4"] = "Métrica"
    ws_sum["B4"] = "Quantidade"
    ws_sum["C4"] = "% do Total"
    for col in ("A4", "B4", "C4"):
        ws_sum[col].font = header_font
        ws_sum[col].fill = blue_fill
        ws_sum[col].alignment = Alignment(horizontal="left" if col == "A4" else "right")

    coverage = with_alert / total if total > 0 else 0
    summary_rows = [
        ("Total de Coletas Master", total, 1.0 if total else 0),
        ("Com Alerta nodata()", with_alert, coverage),
        ("SEM Alerta nodata() (gaps)", without_alert, (1 - coverage) if total else 0),
    ]
    current_row = 5
    for idx, (label, count, pct) in enumerate(summary_rows):
        ws_sum[f"A{current_row}"] = label
        ws_sum[f"B{current_row}"] = count
        ws_sum[f"C{current_row}"] = pct
        ws_sum[f"A{current_row}"].font = bold_font
        ws_sum[f"B{current_row}"].font = regular_font
        ws_sum[f"C{current_row}"].font = regular_font
        ws_sum[f"B{current_row}"].number_format = "#,##0"
        ws_sum[f"C{current_row}"].number_format = "0.0%"
        for col in ("A", "B", "C"):
            ws_sum[f"{col}{current_row}"].border = thin_border
        current_row += 1

    # Breakdown by type
    ws_sum["A9"] = "Tipo de Coleta"
    ws_sum["B9"] = "Total"
    ws_sum["C9"] = "Sem Alerta"
    for col in ("A9", "B9", "C9"):
        ws_sum[col].font = header_font
        ws_sum[col].fill = blue_fill
        ws_sum[col].alignment = Alignment(horizontal="left" if col == "A9" else "right")

    type_total = Counter(r["type"] for r in rows)
    type_gap = Counter(r["type"] for r in rows if not r["has_nodata_alert"])
    current_row = 10
    for idx, (item_type, count) in enumerate(type_total.most_common()):
        ws_sum[f"A{current_row}"] = item_type
        ws_sum[f"B{current_row}"] = count
        ws_sum[f"C{current_row}"] = type_gap.get(item_type, 0)
        use_zebra = (idx % 2 == 1)
        for col in ("A", "B", "C"):
            cell = ws_sum[f"{col}{current_row}"]
            cell.font = regular_font
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill
            if col != "A":
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
        current_row += 1

    # 2. Sheet: Coletas_Master
    ws_det = wb.create_sheet(title="Coletas_Master")
    ws_det.views.sheetView[0].showGridLines = True

    headers = [
        "Template", "ID do Template", "Item", "Chave (Key)", "Tipo de Coleta",
        "Categoria", "Dependentes", "Alerta nodata", "Triggers nodata",
        "Total Triggers", "ID do Item",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="left")

    for idx, r in enumerate(rows):
        row_idx = idx + 2
        row_data = [
            r["template"],
            r["template_id"],
            r["item"],
            r["key"],
            r["type"],
            r["category"],
            r["dependents"],
            "SIM" if r["has_nodata_alert"] else "FALTA",
            r["nodata_triggers"],
            r["trigger_count"],
            r["item_id"],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_det.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            # Highlight the coverage column.
            if col_idx == 8:
                cell.fill = green_fill if r["has_nodata_alert"] else red_fill
                cell.font = bold_font

    for ws in (ws_sum, ws_det):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if ws.title == "Resumo" and cell.row <= 2:
                    continue
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            width = min(max(max_len + 3, 10), 65)
            ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_priorities_excel_in_memory(rows: Sequence[Dict[str, Any]]) -> "io.BytesIO":
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    title_font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)

    blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    priority_fills = {
        "Alta": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "Média": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "Baixa": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    }

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    ws = wb.active
    ws.title = "Ordem_de_Resolucao"
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:H2")
    ws["A1"] = "Ordem de Resolução — Falhas que Mais se Repetem"
    ws["A1"].font = title_font
    ws["A1"].fill = blue_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Ordem", "Prioridade", "Padrão de Erro (Normalizado)", "Itens Afetados",
        "Hosts Afetados", "% do Total", "% Acumulado", "Tipo Dominante",
        "Categoria", "Maior Ofensor (Host)", "Exemplo de Erro",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="left")

    for idx, r in enumerate(rows):
        row_idx = idx + 5
        row_data = [
            r["rank"],
            r["priority"],
            r["pattern"],
            r["count"],
            r["hosts"],
            r["pct"],
            r["cum_pct"],
            r["top_type"],
            r["top_category"],
            r["top_host"],
            r["sample"],
        ]
        use_zebra = (idx % 2 == 1)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if use_zebra and col_idx != 2:
                cell.fill = zebra_fill
            if col_idx in (4, 5):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            if col_idx in (6, 7):
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            if col_idx == 2:
                cell.fill = priority_fills.get(r["priority"], zebra_fill)
                cell.font = bold_font

    ws.freeze_panes = "A5"

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row <= 2:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 3, 10), 70)
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def write_output(rows: Sequence[Dict[str, Any]], fmt: str, output_path: str, report: str = "unsupported") -> None:
    if fmt == "json":
        payload = json.dumps(list(rows), indent=2, ensure_ascii=False)
        if output_path:
            Path(output_path).write_text(payload, encoding="utf-8")
        else:
            print(payload)
        return

    if fmt == "csv":
        if report == "masters":
            fieldnames = [
                "template", "template_id", "item", "key", "type", "category",
                "dependents", "has_nodata_alert", "nodata_triggers",
                "trigger_count", "item_id",
            ]
        elif report == "priorities":
            fieldnames = [
                "rank", "priority", "pattern", "count", "hosts", "pct",
                "cum_pct", "top_type", "top_category", "top_host", "sample",
            ]
        else:
            fieldnames = ["host", "host_id", "item", "key", "type", "category", "error", "item_id"]
        if output_path:
            handle = Path(output_path).open("w", newline="", encoding="utf-8")
            close_handle = True
        else:
            handle = sys.stdout
            close_handle = False

        try:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow({name: row.get(name, "") for name in fieldnames})
        finally:
            if close_handle:
                handle.close()
        return

    if fmt == "excel":
        if not output_path:
            raise SystemExit("Argumento --output e obrigatorio para o formato 'excel'.")
        try:
            if report == "masters":
                excel_io = generate_masters_excel_in_memory(rows)
            elif report == "priorities":
                excel_io = generate_priorities_excel_in_memory(rows)
            else:
                excel_io = generate_excel_in_memory(rows)
            Path(output_path).write_bytes(excel_io.getbuffer())
            print(f"Relatorio Excel gerado com sucesso em: {output_path}")
        except Exception as e:
            raise SystemExit(f"Erro ao salvar arquivo Excel: {e}")
        return

    if report == "masters":
        table = render_masters_table(rows)
    elif report == "priorities":
        table = render_priorities_table(rows)
    else:
        table = render_table(rows)
    if output_path:
        Path(output_path).write_text(table + "\n", encoding="utf-8")
    else:
        print(table)


_cache: Dict[str, Any] = {
    "data": None,
    "timestamp": 0.0
}

_masters_cache: Dict[str, Any] = {
    "data": None,
    "timestamp": 0.0
}


def collect_data(settings: Settings, days: int = 0) -> List[Dict[str, Any]]:
    # We ignore days for unsupported items since items don't have active problem timers
    api = build_api(settings)

    # Fetch active (enabled=0) items that are unsupported (state=1)
    items = api.item.get(
        filter={"state": 1, "status": 0},
        output=["itemid", "name", "key_", "state", "error", "type"],
        selectHosts=["hostid", "host", "name"]
    )
    
    # Extract unique host IDs
    host_ids = set()
    for item in items:
        hosts = item.get("hosts", [])
        if hosts:
            host_ids.add(hosts[0]["hostid"])
            
    # Fetch hosts with their groups
    host_groups_map = {}
    if host_ids:
        host_ids_list = sorted(list(host_ids))
        for i in range(0, len(host_ids_list), 200):
            chunk = host_ids_list[i:i + 200]
            hosts_result = api.host.get(
                hostids=chunk,
                output=["hostid"],
                selectGroups=["name"]
            )
            for h in hosts_result:
                groups = [g["name"] for g in h.get("groups", []) or [] if g.get("name")]
                host_groups_map[h["hostid"]] = groups

    rows: List[Dict[str, Any]] = []
    for item in items:
        hosts = item.get("hosts", [])
        if hosts:
            host_name = hosts[0].get("name") or hosts[0].get("host") or "-"
            host_id = hosts[0].get("hostid") or "-"
        else:
            host_name = "-"
            host_id = "-"

        # Classify host based on group names
        groups = host_groups_map.get(host_id, [])
        category = classify_host(groups)

        rows.append({
            "host": host_name,
            "host_id": host_id,
            "item": item.get("name") or "-",
            "key": item.get("key_") or "-",
            "type": get_item_type_label(item.get("type")),
            "type_code": str(item.get("type", "")),
            "category": category,
            "error": item.get("error") or "Unsupported item (no error description)",
            "item_id": item.get("itemid", "-")
        })

    # Sort by host name, then item name
    rows.sort(key=lambda r: (r["host"].lower(), r["item"].lower()))
    return rows


def collect_master_items(settings: Settings) -> List[Dict[str, Any]]:
    """Find template-level master collection items and their nodata() coverage.

    A "master collection" here is an item that (a) is of a bulk-collection type
    (Script, External check, HTTP agent, DB monitor, SSH, Telnet, Calculated)
    AND (b) is referenced as master_itemid by at least one dependent item.

    The scope is per *template*, not per host: we only look at items defined
    directly on templates (templated=True, inherited=False), so each master is
    reported once regardless of how many hosts inherit it. For each master we
    inspect its triggers and flag whether any uses nodata() (i.e. there is an
    alert for collection failure).
    """
    from collections import defaultdict

    api = build_api(settings)

    # Step 1: discover masters by scanning dependent items defined directly on
    # templates (not inherited copies), so we stay at the template level.
    dependent_items = api.item.get(
        templated=True,
        inherited=False,
        filter={"type": int(DEPENDENT_ITEM_TYPE)},
        output=["itemid", "master_itemid"],
    )

    dependents_count: Dict[str, int] = defaultdict(int)
    for dep in dependent_items:
        master_id = dep.get("master_itemid")
        if master_id and master_id != "0":
            dependents_count[master_id] += 1

    master_ids = sorted(dependents_count.keys())
    if not master_ids:
        return []

    # Step 2: fetch the master items themselves, along with their triggers.
    # selectHosts on a template item returns the owning template (status == 3).
    master_items: List[Dict[str, Any]] = []
    for i in range(0, len(master_ids), 500):
        chunk = master_ids[i:i + 500]
        result = api.item.get(
            itemids=chunk,
            output=["itemid", "name", "key_", "type", "status", "state"],
            selectHosts=["hostid", "host", "name", "status"],
            selectTriggers=["triggerid", "description"],
        )
        master_items.extend(result)

    # Step 2b: resolve trigger functions so we can detect nodata() precisely.
    # The trigger.get "expression" field returns functionid references
    # (e.g. "{18175}=1"), so we inspect functions instead: a master is covered
    # when a trigger has a nodata() function pointing at the master's own item.
    all_trigger_ids = sorted({
        t.get("triggerid")
        for item in master_items
        for t in (item.get("triggers") or [])
        if t.get("triggerid")
    })

    # trigger_id -> set of itemids that have a nodata() function in that trigger.
    nodata_by_trigger: Dict[str, set] = {}
    for i in range(0, len(all_trigger_ids), 1000):
        chunk = all_trigger_ids[i:i + 1000]
        triggers = api.trigger.get(
            triggerids=chunk,
            output=["triggerid"],
            selectFunctions="extend",
        )
        for trg in triggers:
            nodata_items = {
                fn.get("itemid")
                for fn in (trg.get("functions") or [])
                if (fn.get("function") or "").lower() == NODATA_KEYWORD
            }
            if nodata_items:
                nodata_by_trigger[trg["triggerid"]] = nodata_items

    # Step 3: resolve template groups for classification. Templates are not
    # returned by host.get, so we query template.get with selectGroups.
    template_ids = set()
    for item in master_items:
        hosts = item.get("hosts", [])
        if hosts:
            template_ids.add(hosts[0]["hostid"])

    template_groups_map: Dict[str, List[str]] = {}
    if template_ids:
        template_ids_list = sorted(template_ids)
        for i in range(0, len(template_ids_list), 200):
            chunk = template_ids_list[i:i + 200]
            templates_result = api.template.get(
                templateids=chunk,
                output=["templateid"],
                selectGroups=["name"],
            )
            for t in templates_result:
                groups = [g["name"] for g in t.get("groups", []) or [] if g.get("name")]
                template_groups_map[t["templateid"]] = groups

    rows: List[Dict[str, Any]] = []
    for item in master_items:
        type_code = str(item.get("type", ""))
        # Keep only bulk-collection types (the "Ambos combinados" definition).
        if type_code not in MASTER_COLLECTION_TYPES:
            continue

        hosts = item.get("hosts", [])
        if hosts:
            template_name = hosts[0].get("name") or hosts[0].get("host") or "-"
            template_id = hosts[0].get("hostid") or "-"
        else:
            template_name = "-"
            template_id = "-"

        groups = template_groups_map.get(template_id, [])
        category = classify_host(groups)

        item_id = item["itemid"]
        triggers = item.get("triggers", []) or []
        # A trigger is a nodata alert for THIS master when it has a nodata()
        # function pointing at the master's own itemid.
        nodata_triggers = [
            t for t in triggers
            if item_id in nodata_by_trigger.get(t.get("triggerid"), set())
        ]
        has_nodata = bool(nodata_triggers)

        rows.append({
            "template": template_name,
            "template_id": template_id,
            "item": item.get("name") or "-",
            "key": item.get("key_") or "-",
            "type": get_item_type_label(item.get("type")),
            "type_code": type_code,
            "category": category,
            "dependents": dependents_count.get(item_id, 0),
            "has_nodata_alert": has_nodata,
            "nodata_triggers": "; ".join(
                t.get("description", "") for t in nodata_triggers
            ) or "-",
            "trigger_count": len(triggers),
            "item_id": item.get("itemid", "-"),
        })

    # Gaps first (items without a nodata alert), then by template and item name.
    rows.sort(key=lambda r: (r["has_nodata_alert"], r["template"].lower(), r["item"].lower()))
    return rows


def aggregate_priorities(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Group unsupported items by normalized error to build a resolution order.

    Each group is ranked by impact (number of affected items, then breadth in
    distinct hosts); rank 1 is what should be fixed first. The priority badge
    combines the pattern's share of all failures with its rank, so the top
    recurring failures are always flagged for focus even when one pattern
    dominates the total:
      - Alta:  >= 5% of all failures, or among the top 3 patterns
      - Média: >= 1% of all failures, or among the top 10 patterns
      - Baixa: everything else (the long tail)
    The cumulative percentage (cum_pct) is kept as Pareto context.
    """
    from collections import Counter, defaultdict

    total = len(rows)

    groups: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
        "count": 0,
        "hosts": Counter(),
        "types": Counter(),
        "categories": Counter(),
        "sample": "",
    })

    for r in rows:
        pattern = normalize_error(r.get("error", ""))
        group = groups[pattern]
        group["count"] += 1
        host = r.get("host")
        if host and host != "-":
            group["hosts"][host] += 1
        group["types"][r.get("type", "-")] += 1
        group["categories"][r.get("category", "All")] += 1
        if not group["sample"]:
            group["sample"] = r.get("error", "") or ""

    items: List[Dict[str, Any]] = []
    for pattern, group in groups.items():
        items.append({
            "pattern": pattern,
            "count": group["count"],
            "hosts": len(group["hosts"]),
            "top_host": group["hosts"].most_common(1)[0][0] if group["hosts"] else "-",
            "top_type": group["types"].most_common(1)[0][0] if group["types"] else "-",
            "top_category": group["categories"].most_common(1)[0][0] if group["categories"] else "All",
            "sample": group["sample"],
        })

    # Most repeated first; ties broken by how many hosts are affected (breadth).
    items.sort(key=lambda x: (-x["count"], -x["hosts"], x["pattern"]))

    cumulative = 0
    for idx, item in enumerate(items):
        cumulative += item["count"]
        item["rank"] = idx + 1
        item["pct"] = item["count"] / total if total else 0
        item["cum_pct"] = cumulative / total if total else 0
        if item["pct"] >= 0.05 or item["rank"] <= 3:
            item["priority"] = "Alta"
        elif item["pct"] >= 0.01 or item["rank"] <= 10:
            item["priority"] = "Média"
        else:
            item["priority"] = "Baixa"

    return items


def run_server(settings: Settings, days: int, port: int) -> None:
    from flask import Flask, jsonify, render_template, request, send_file

    app = Flask(__name__, template_folder="templates")

    # Disable default Werkzeug logs to keep the console clean
    import logging
    log = logging.getLogger('werkzeug')
    log.setLevel(logging.WARNING)

    @app.route("/")
    def index():
        return render_template("index.html")

    @app.route("/api/problems")
    def api_problems():
        force = request.args.get("force", "false").lower() == "true"
        current_time = time.time()
        
        try:
            if not force and _cache["data"] is not None and (current_time - _cache["timestamp"]) < settings.cache_ttl:
                return jsonify(_cache["data"])

            data = collect_data(settings, days=days)
            _cache["data"] = data
            _cache["timestamp"] = current_time
            return jsonify(data)
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    def _get_unsupported_cached(force=False):
        current_time = time.time()
        if (not force and _cache["data"] is not None
                and (current_time - _cache["timestamp"]) < settings.cache_ttl):
            return _cache["data"]
        data = collect_data(settings, days=days)
        _cache["data"] = data
        _cache["timestamp"] = current_time
        return data

    @app.route("/api/priorities")
    def api_priorities():
        force = request.args.get("force", "false").lower() == "true"
        try:
            return jsonify(aggregate_priorities(_get_unsupported_cached(force)))
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/export/priorities/excel")
    def api_export_priorities_excel():
        try:
            rows = aggregate_priorities(_get_unsupported_cached())
            excel_io = generate_priorities_excel_in_memory(rows)
            return send_file(
                excel_io,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="ordem_de_resolucao.xlsx",
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/masters")
    def api_masters():
        force = request.args.get("force", "false").lower() == "true"
        current_time = time.time()

        try:
            if (not force and _masters_cache["data"] is not None
                    and (current_time - _masters_cache["timestamp"]) < settings.cache_ttl):
                return jsonify(_masters_cache["data"])

            data = collect_master_items(settings)
            _masters_cache["data"] = data
            _masters_cache["timestamp"] = current_time
            return jsonify(data)
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/export/masters/excel")
    def api_export_masters_excel():
        current_time = time.time()
        try:
            if (_masters_cache["data"] is not None
                    and (current_time - _masters_cache["timestamp"]) < settings.cache_ttl):
                data = _masters_cache["data"]
            else:
                data = collect_master_items(settings)
                _masters_cache["data"] = data
                _masters_cache["timestamp"] = current_time

            excel_io = generate_masters_excel_in_memory(data)
            return send_file(
                excel_io,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="relatorio_coletas_master.xlsx",
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/export/excel")
    def api_export_excel():
        current_time = time.time()
        try:
            # Re-use cached data if valid, otherwise fetch fresh
            if _cache["data"] is not None and (current_time - _cache["timestamp"]) < settings.cache_ttl:
                data = _cache["data"]
            else:
                data = collect_data(settings, days=days)
                _cache["data"] = data
                _cache["timestamp"] = current_time

            excel_io = generate_excel_in_memory(data)
            return send_file(
                excel_io,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="relatorio_coleta_zabbix.xlsx"
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    print("==================================================")
    print(" Painel de Monitoramento Zabbix Iniciado!")
    print(f" Acesse em seu navegador: http://localhost:{port}")
    print(" Pressione Ctrl+C para encerrar o servidor.")
    print("==================================================")

    app.run(host="0.0.0.0", port=port, debug=False)


def main() -> int:
    args = parse_args()
    settings = load_settings(args)

    if args.serve:
        run_server(settings, args.days, args.port)
        return 0

    if args.report == "masters":
        rows = collect_master_items(settings)
        write_output(rows, args.format, args.output, report="masters")
        if rows:
            without_alert = sum(1 for r in rows if not r["has_nodata_alert"])
            print(
                f"Coletas master: {len(rows)} | sem alerta nodata(): {without_alert}",
                file=sys.stderr,
            )
        else:
            print("Nenhum item master de coleta foi identificado.", file=sys.stderr)
        return 0

    if args.report == "priorities":
        failures = collect_data(settings, days=args.days)
        rows = aggregate_priorities(failures)
        write_output(rows, args.format, args.output, report="priorities")
        if rows:
            print(
                f"Padrões de falha: {len(rows)} | total de itens: {len(failures)} | "
                f"foco (Alta): {sum(1 for r in rows if r['priority'] == 'Alta')}",
                file=sys.stderr,
            )
        else:
            print("Nenhuma falha encontrada para priorizar.", file=sys.stderr)
        return 0

    rows = collect_data(settings, days=args.days)
    write_output(rows, args.format, args.output, report="unsupported")

    if rows:
        print(f"Total de itens não suportados: {len(rows)}", file=sys.stderr)
    else:
        print("Nenhum item não suportado foi identificado.", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())